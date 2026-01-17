#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel数据筛选工具
支持多条件筛选，可将符合条件的数据追加到指定Sheet
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import os
import sys


class ExcelFilterTool:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel数据筛选工具")
        self.root.geometry("900x700")

        self.file_path = None
        self.workbook = None
        self.sheet_names = []
        self.headers = []
        self.conditions = []

        self.setup_ui()

    def setup_ui(self):
        """设置UI界面"""
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # 配置网格权重
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)

        row = 0

        # 文件选择部分
        ttk.Label(main_frame, text="选择Excel文件:").grid(row=row, column=0, sticky=tk.W, pady=5)
        row += 1

        file_frame = ttk.Frame(main_frame)
        file_frame.grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        main_frame.columnconfigure(1, weight=1)

        self.file_entry = ttk.Entry(file_frame, width=50)
        self.file_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

        ttk.Button(file_frame, text="浏览", command=self.select_file).pack(side=tk.LEFT)
        row += 1

        # Sheet选择部分
        ttk.Label(main_frame, text="选择要处理的Sheet:").grid(row=row, column=0, sticky=tk.W, pady=5)
        row += 1

        self.sheet_combobox = ttk.Combobox(main_frame, width=40, state='readonly')
        self.sheet_combobox.grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        self.sheet_combobox.bind('<<ComboboxSelected>>', self.on_sheet_selected)
        row += 1

        # 分隔线
        ttk.Separator(main_frame, orient='horizontal').grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        row += 1

        # 条件设置部分
        ttk.Label(main_frame, text="筛选条件设置", font=('Arial', 12, 'bold')).grid(row=row, column=0, columnspan=3, pady=5)
        row += 1

        # 条件列表框架
        conditions_frame = ttk.LabelFrame(main_frame, text="已添加的条件", padding="5")
        conditions_frame.grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        main_frame.rowconfigure(row, weight=1)

        # 条件列表树形视图
        columns = ('列名', '运算符', '值', '逻辑关系')
        self.conditions_tree = ttk.Treeview(conditions_frame, columns=columns, show='headings', height=6)

        for col in columns:
            self.conditions_tree.heading(col, text=col)
            self.conditions_tree.column(col, width=150, anchor=tk.CENTER)

        # 滚动条
        scrollbar = ttk.Scrollbar(conditions_frame, orient=tk.VERTICAL, command=self.conditions_tree.yview)
        self.conditions_tree.configure(yscrollcommand=scrollbar.set)

        self.conditions_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))

        conditions_frame.columnconfigure(0, weight=1)
        conditions_frame.rowconfigure(0, weight=1)

        # 条件操作按钮
        btn_frame = ttk.Frame(conditions_frame)
        btn_frame.grid(row=1, column=0, columnspan=2, pady=5)

        ttk.Button(btn_frame, text="添加条件", command=self.add_condition).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="删除条件", command=self.delete_condition).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="清空条件", command=self.clear_conditions).pack(side=tk.LEFT, padx=5)
        row += 1

        # 目标Sheet设置
        ttk.Label(main_frame, text="目标Sheet名称:").grid(row=row, column=0, sticky=tk.W, pady=5)
        row += 1

        target_frame = ttk.Frame(main_frame)
        target_frame.grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)

        self.target_sheet_entry = ttk.Entry(target_frame, width=40)
        self.target_sheet_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.target_sheet_entry.insert(0, "筛选结果")

        ttk.Label(target_frame, text="(留空则创建新Sheet)").pack(side=tk.LEFT, padx=5)
        row += 1

        # 执行按钮
        ttk.Separator(main_frame, orient='horizontal').grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        row += 1

        btn_execute_frame = ttk.Frame(main_frame)
        btn_execute_frame.grid(row=row, column=0, columnspan=3, pady=10)

        ttk.Button(btn_execute_frame, text="执行筛选", command=self.execute_filter, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_execute_frame, text="保存结果", command=self.save_result, width=20).pack(side=tk.LEFT, padx=5)

        # 状态栏
        row += 1
        self.status_label = ttk.Label(main_frame, text="就绪", relief=tk.SUNKEN, anchor=tk.W)
        self.status_label.grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(10, 0))

        self.result_workbook = None
        self.result_file_path = None

    def select_file(self):
        """选择Excel文件"""
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )

        if file_path:
            self.file_path = file_path
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, file_path)

            try:
                # 加载工作簿
                self.workbook = load_workbook(file_path, data_only=True)
                self.sheet_names = self.workbook.sheetnames

                # 更新Sheet下拉框
                self.sheet_combobox['values'] = self.sheet_names
                if self.sheet_names:
                    self.sheet_combobox.current(0)

                self.update_status(f"已加载文件: {os.path.basename(file_path)}")
            except Exception as e:
                messagebox.showerror("错误", f"加载文件失败: {str(e)}")
                self.update_status("加载文件失败")

    def on_sheet_selected(self, event):
        """当选择Sheet时，读取表头"""
        selected_sheet = self.sheet_combobox.get()
        if not selected_sheet or not self.workbook:
            return

        try:
            sheet = self.workbook[selected_sheet]
            self.headers = []

            # 读取第一行作为表头
            for cell in sheet[1]:
                if cell.value:
                    self.headers.append(str(cell.value))
                else:
                    break

            self.update_status(f"已加载Sheet '{selected_sheet}'，共 {len(self.headers)} 列")
        except Exception as e:
            messagebox.showerror("错误", f"读取Sheet失败: {str(e)}")

    def add_condition(self):
        """添加筛选条件"""
        if not self.headers:
            messagebox.showwarning("警告", "请先选择Excel文件和Sheet")
            return

        # 创建条件设置对话框
        dialog = ConditionDialog(self.root, self.headers, self.conditions)
        self.root.wait_window(dialog.dialog)

        if dialog.result:
            condition = dialog.result
            self.conditions.append(condition)

            # 更新条件列表
            logic_text = condition['logic'] if condition['logic'] else ""
            self.conditions_tree.insert('', tk.END, values=(
                condition['column'],
                condition['operator'],
                condition['value'],
                logic_text
            ))

    def delete_condition(self):
        """删除选中的条件"""
        selected = self.conditions_tree.selection()
        if not selected:
            messagebox.showwarning("警告", "请先选择要删除的条件")
            return

        index = self.conditions_tree.index(selected[0])
        del self.conditions[index]
        self.conditions_tree.delete(selected[0])

    def clear_conditions(self):
        """清空所有条件"""
        if not self.conditions:
            return

        if messagebox.askyesno("确认", "确定要清空所有条件吗？"):
            self.conditions.clear()
            for item in self.conditions_tree.get_children():
                self.conditions_tree.delete(item)

    def execute_filter(self):
        """执行筛选"""
        if not self.file_path or not self.workbook:
            messagebox.showwarning("警告", "请先选择Excel文件")
            return

        if not self.conditions:
            messagebox.showwarning("警告", "请至少添加一个筛选条件")
            return

        source_sheet_name = self.sheet_combobox.get()
        if not source_sheet_name:
            messagebox.showwarning("警告", "请选择要处理的Sheet")
            return

        try:
            source_sheet = self.workbook[source_sheet_name]

            # 找到表头行
            header_row = 1
            header_indices = {}
            for idx, cell in enumerate(source_sheet[header_row], start=1):
                if cell.value:
                    header_indices[str(cell.value)] = idx
                else:
                    break

            # 筛选数据
            filtered_rows = []
            filtered_rows.append([cell.value for cell in source_sheet[header_row]])  # 添加表头

            for row in source_sheet.iter_rows(min_row=header_row + 1):
                if self.check_conditions(row, header_indices):
                    filtered_rows.append([cell.value for cell in row])

            if len(filtered_rows) == 1:
                messagebox.showinfo("提示", "没有找到符合条件的数据")
                return

            # 创建或打开目标Sheet
            target_sheet_name = self.target_sheet_entry.get().strip() or "筛选结果"

            # 检查目标Sheet是否存在
            if target_sheet_name in self.workbook.sheetnames:
                # 追加到现有Sheet
                target_sheet = self.workbook[target_sheet_name]
                # 找到最后一行
                last_row = target_sheet.max_row

                # 如果是空Sheet或者只有表头，从第一行开始
                if last_row == 1:
                    start_row = 1
                else:
                    start_row = last_row + 1

                # 写入数据（如果表头已存在，跳过表头）
                data_start = 0 if last_row == 1 else 1
                for row_idx, row_data in enumerate(filtered_rows[data_start:], start=start_row):
                    for col_idx, value in enumerate(row_data, start=1):
                        target_sheet.cell(row=row_idx, column=col_idx, value=value)

            else:
                # 创建新Sheet
                target_sheet = self.workbook.create_sheet(title=target_sheet_name)
                for row_idx, row_data in enumerate(filtered_rows, start=1):
                    for col_idx, value in enumerate(row_data, start=1):
                        target_sheet.cell(row=row_idx, column=col_idx, value=value)

            # 自动保存到原文件
            try:
                self.workbook.save(self.file_path)
                self.result_workbook = None  # 已保存，不需要再保存
                self.update_status(f"筛选完成！共找到 {len(filtered_rows) - 1} 行数据，已保存到文件")
                messagebox.showinfo("成功", f"筛选完成！\n共找到 {len(filtered_rows) - 1} 行符合条件的数据\n已自动保存到原文件\n目标Sheet: {target_sheet_name}")
            except Exception as save_error:
                # 如果保存失败，尝试另存为
                messagebox.showwarning("保存失败", f"无法保存到原文件: {str(save_error)}\n\n请点击'保存结果'按钮另存为新文件")
                self.result_workbook = self.workbook
                self.update_status(f"筛选完成！共找到 {len(filtered_rows) - 1} 行数据，请保存结果")

        except Exception as e:
            messagebox.showerror("错误", f"执行筛选失败: {str(e)}")
            self.update_status("执行失败")

    def check_conditions(self, row, header_indices):
        """检查行数据是否符合所有条件"""
        for condition in self.conditions:
            column_name = condition['column']
            operator = condition['operator']
            condition_value = condition['value']

            if column_name not in header_indices:
                continue

            col_index = header_indices[column_name]
            cell_value = row[col_index - 1].value

            # 处理空值
            if cell_value is None:
                cell_value = ""
            else:
                try:
                    cell_value = str(cell_value)
                except:
                    cell_value = str(cell_value)

            condition_value = str(condition_value)

            # 根据运算符判断
            result = False
            try:
                if operator == "=":
                    result = (cell_value == condition_value)
                elif operator == "!=":
                    result = (cell_value != condition_value)
                elif operator == ">=":
                    try:
                        result = (float(cell_value) >= float(condition_value))
                    except:
                        result = (cell_value >= condition_value)
                elif operator == "<=":
                    try:
                        result = (float(cell_value) <= float(condition_value))
                    except:
                        result = (cell_value <= condition_value)
                elif operator == ">":
                    try:
                        result = (float(cell_value) > float(condition_value))
                    except:
                        result = (cell_value > condition_value)
                elif operator == "<":
                    try:
                        result = (float(cell_value) < float(condition_value))
                    except:
                        result = (cell_value < condition_value)
                elif operator == "包含":
                    result = (condition_value in cell_value)
                elif operator == "不包含":
                    result = (condition_value not in cell_value)
                else:
                    result = True
            except:
                result = False

            # 根据逻辑关系决定是否继续检查
            if condition['logic'] == "或":
                if result:
                    return True
            elif condition['logic'] == "与":
                if not result:
                    return False
            else:
                # 最后一个条件
                if not result:
                    return False

        return True

    def save_result(self):
        """保存结果到新文件"""
        if not self.result_workbook:
            messagebox.showwarning("警告", "请先执行筛选")
            return

        # 生成默认文件名
        default_name = "筛选结果_" + os.path.basename(self.file_path)

        file_path = filedialog.asksaveasfilename(
            title="保存筛选结果",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )

        if file_path:
            try:
                self.result_workbook.save(file_path)
                self.result_file_path = file_path
                self.update_status(f"已保存到: {os.path.basename(file_path)}")
                messagebox.showinfo("成功", f"文件已保存到:\n{file_path}")
            except Exception as e:
                messagebox.showerror("错误", f"保存文件失败: {str(e)}")

    def update_status(self, message):
        """更新状态栏"""
        self.status_label.config(text=message)


class ConditionDialog:
    """条件设置对话框"""

    def __init__(self, parent, headers, existing_conditions):
        self.result = None
        self.headers = headers
        self.existing_conditions = existing_conditions

        self.dialog = tk.Toplevel(parent)
        self.dialog.title("添加筛选条件")
        self.dialog.geometry("500x350")
        self.dialog.transient(parent)
        self.dialog.grab_set()

        self.setup_ui()

        # 居中显示
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - (self.dialog.winfo_width() // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (self.dialog.winfo_height() // 2)
        self.dialog.geometry(f"+{x}+{y}")

    def setup_ui(self):
        """设置对话框UI"""
        main_frame = ttk.Frame(self.dialog, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        self.dialog.columnconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)

        row = 0

        # 列名选择
        ttk.Label(main_frame, text="选择列:").grid(row=row, column=0, sticky=tk.W, pady=10)
        row += 1

        self.column_var = tk.StringVar()
        self.column_combobox = ttk.Combobox(main_frame, textvariable=self.column_var, values=self.headers, state='readonly')
        self.column_combobox.grid(row=row, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        if self.headers:
            self.column_combobox.current(0)
        row += 1

        # 运算符选择
        ttk.Label(main_frame, text="运算符:").grid(row=row, column=0, sticky=tk.W, pady=10)
        row += 1

        operators = ["=", "!=", ">=", "<=", ">", "<", "包含", "不包含"]
        self.operator_var = tk.StringVar(value="=")
        operator_combobox = ttk.Combobox(main_frame, textvariable=self.operator_var, values=operators, state='readonly')
        operator_combobox.grid(row=row, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        row += 1

        # 值输入
        ttk.Label(main_frame, text="比较值:").grid(row=row, column=0, sticky=tk.W, pady=10)
        row += 1

        self.value_var = tk.StringVar()
        self.value_entry = ttk.Entry(main_frame, textvariable=self.value_var, width=40)
        self.value_entry.grid(row=row, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        row += 1

        # 逻辑关系
        ttk.Label(main_frame, text="与下一条件的关系:").grid(row=row, column=0, sticky=tk.W, pady=10)
        row += 1

        # 总是添加"无"选项，对于第一个条件默认选择"无"
        self.logic_var = tk.StringVar(value="")
        logic_frame = ttk.Frame(main_frame)
        logic_frame.grid(row=row, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)

        ttk.Radiobutton(logic_frame, text="与 (AND)", variable=self.logic_var, value="与").pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(logic_frame, text="或 (OR)", variable=self.logic_var, value="或").pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(logic_frame, text="无 (最后一个条件)", variable=self.logic_var, value="").pack(side=tk.LEFT, padx=10)

        # 如果已有条件，默认选择"与"
        if self.existing_conditions:
            self.logic_var.set("与")
        row += 1

        # 按钮框架
        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=row, column=0, columnspan=2, pady=20)

        ttk.Button(btn_frame, text="确定", command=self.on_ok, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="取消", command=self.on_cancel, width=10).pack(side=tk.LEFT, padx=5)

    def on_ok(self):
        """确定按钮"""
        column = self.column_var.get()
        operator = self.operator_var.get()
        value = self.value_var.get()
        logic = self.logic_var.get()

        if not column:
            messagebox.showwarning("警告", "请选择列")
            return

        if not value:
            messagebox.showwarning("警告", "请输入比较值")
            return

        self.result = {
            'column': column,
            'operator': operator,
            'value': value,
            'logic': logic
        }
        self.dialog.destroy()

    def on_cancel(self):
        """取消按钮"""
        self.dialog.destroy()


def main():
    root = tk.Tk()
    app = ExcelFilterTool(root)
    root.mainloop()


if __name__ == "__main__":
    main()
